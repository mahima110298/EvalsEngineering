#!/usr/bin/env python3
"""
tester_agent.py — 3-simulation evaluation harness for todo_agent.py

Runs all 100 test cases in parallel per simulation (capped at 20 workers),
prints per-simulation accuracy tables, and exports failures to Excel.

Setup:  bash setup_local.sh
Run:    python3 tester_agent.py
"""

import json
import sys
import time
import concurrent.futures
from datetime import datetime

import ollama
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from test_cases_full import TEST_CASES as _ALL_TESTS

# 20 representative tests — one per logical pattern, covering all 8 categories
_SELECTED_IDS = {
    "S1","S3","S5","S7",          # simple: add, complete, priority edit, date edit
    "A1","A6","A9",               # ambiguity: 3-match clarify, ordinal complete, exact-match confirm
    "C1","C5","C10",              # correction: 2-turn date fix, 2-turn redirect, in-utterance fix
    "D2","D9","D10",              # deletion: bulk confirm, in-utterance cancel, never-mind cancel
    "N1","N5",                    # noisy: ASR typo, filler word
    "M1","M9",                    # multi-intent: two adds, bulk complete
    "O1",                         # over-confirmation: add must NOT confirm
    "H1","H2",                    # natural speech: colloquial add, "I finished X"
}
TEST_CASES = [t for t in _ALL_TESTS if t["id"] in _SELECTED_IDS]
assert len(TEST_CASES) == 20, f"Expected 20, got {len(TEST_CASES)}"

MODEL        = "llama3.2:latest"   # 3.2B — already on disk, ~3x faster than gemma4
SIMULATIONS  = 3
MAX_WORKERS  = 30

# ---------------------------------------------------------------------------
# Tool schemas
# ---------------------------------------------------------------------------
TOOLS = [
    {"type":"function","function":{"name":"add","description":"Add a new todo item.",
     "parameters":{"type":"object","properties":{
         "title":{"type":"string"},"description":{"type":"string"},
         "due_date":{"type":"string"},"priority":{"type":"string","enum":["low","normal","high"]}},
     "required":["title"]}}},
    {"type":"function","function":{"name":"edit","description":"Edit an existing todo by its numeric ID.",
     "parameters":{"type":"object","properties":{
         "id":{"type":"integer"},"title":{"type":"string"},"description":{"type":"string"},
         "due_date":{"type":"string"},"priority":{"type":"string","enum":["low","normal","high"]}},
     "required":["id"]}}},
    {"type":"function","function":{"name":"delete","description":"Delete a todo by ID. ALWAYS call confirm first.",
     "parameters":{"type":"object","properties":{"id":{"type":"integer"}},"required":["id"]}}},
    {"type":"function","function":{"name":"complete","description":"Mark a todo complete by ID.",
     "parameters":{"type":"object","properties":{"id":{"type":"integer"}},"required":["id"]}}},
    {"type":"function","function":{"name":"ask_clarification",
     "description":"Ask a clarifying question when the request is ambiguous.",
     "parameters":{"type":"object","properties":{"question":{"type":"string"}},"required":["question"]}}},
    {"type":"function","function":{"name":"confirm",
     "description":"Ask user to confirm a destructive or bulk action. Required before any delete.",
     "parameters":{"type":"object","properties":{"action_description":{"type":"string"}},"required":["action_description"]}}},
]

# ---------------------------------------------------------------------------
# Isolated agent runner — one instance per test, zero shared state
# ---------------------------------------------------------------------------

class TodoAgentRunner:
    def __init__(self, initial_todos: list[dict]):
        self.todos: list[dict] = []
        self.next_id = 1
        for t in initial_todos:
            self.todos.append({
                "id": self.next_id, "title": t.get("title",""),
                "description": t.get("description",""), "completed": t.get("completed", False),
                "due_date": t.get("due_date",""), "priority": t.get("priority","normal"),
            })
            self.next_id += 1

    def _display(self) -> str:
        if not self.todos:
            return "No todos yet."
        lines = []
        for i, t in enumerate(self.todos, 1):
            status = "✓" if t["completed"] else "○"
            meta = []
            if t.get("due_date"):   meta.append(f"due:{t['due_date']}")
            if t.get("priority","normal") != "normal": meta.append(f"priority:{t['priority']}")
            if t["completed"]:      meta.append("completed")
            suffix = f" ({', '.join(meta)})" if meta else ""
            lines.append(f"  {i}. [{t['id']}] {status} {t['title']}{suffix}")
        return "\n".join(lines)

    def _add(self, title="", description="", due_date="", priority="normal", **_):
        todo = {"id":self.next_id,"title":title,"description":description,
                "completed":False,"due_date":due_date,"priority":priority}
        self.todos.append(todo); self.next_id += 1
        return {"success":True,"id":todo["id"],"message":f"Added: '{title}'"}

    def _edit(self, id=0, title=None, description=None, due_date=None, priority=None, **_):
        for t in self.todos:
            if t["id"] == id:
                if title       is not None: t["title"]       = title
                if description is not None: t["description"] = description
                if due_date    is not None: t["due_date"]    = due_date
                if priority    is not None: t["priority"]    = priority
                return {"success":True,"message":f"Updated #{id}"}
        return {"success":False,"message":f"#{id} not found"}

    def _delete(self, id=0, **_):
        for i, t in enumerate(self.todos):
            if t["id"] == id:
                removed = self.todos.pop(i)
                return {"success":True,"message":f"Deleted: '{removed['title']}'"}
        return {"success":False,"message":f"#{id} not found"}

    def _complete(self, id=0, **_):
        for t in self.todos:
            if t["id"] == id:
                t["completed"] = True
                return {"success":True,"message":f"Completed: '{t['title']}'"}
        return {"success":False,"message":f"#{id} not found"}

    def _ask_clarification(self, question="", **_):
        return {"clarification":"Tester: unclear, use best judgment."}

    def _confirm(self, action_description="", **_):
        return {"confirmed":False,"message":"Tester: cancelled to prevent real changes."}

    def _exec(self, name: str, args: dict) -> dict:
        return {"add":self._add,"edit":self._edit,"delete":self._delete,"complete":self._complete,
                "ask_clarification":self._ask_clarification,"confirm":self._confirm
               }.get(name, lambda **_: {"error":f"Unknown: {name}"})(**args)

    def run_turn(self, user_message: str) -> dict:
        tool_calls: list[dict] = []
        system = (
            f"Todo list assistant. Todos:\n{self._display()}\n\n"
            "Rules: call confirm before delete. Call ask_clarification tool (never plain text) "
            "when target is ambiguous. Act directly for safe commands—do not over-confirm. "
            "Infer misspelled/noisy input rather than asking. Never fabricate IDs."
        )
        messages = [{"role":"system","content":system},{"role":"user","content":user_message}]
        final_text = ""
        while True:
            resp = ollama.chat(model=MODEL, messages=messages, tools=TOOLS)
            msg = resp.message
            messages.append(msg)
            if not msg.tool_calls:
                final_text = msg.content or ""
                break
            for tc in msg.tool_calls:
                name = tc.function.name
                args = (tc.function.arguments if isinstance(tc.function.arguments, dict)
                        else json.loads(tc.function.arguments))
                tool_calls.append({"name":name,"args":args})
                messages.append({"role":"tool","content":json.dumps(self._exec(name,args)),"name":name})
        return {"user":user_message,"tool_calls":tool_calls,"response":final_text,
                "todos_after":[t.copy() for t in self.todos]}

# ---------------------------------------------------------------------------
# Judge
# ---------------------------------------------------------------------------

_A2T: dict[str, str | None] = {
    "ADD_TASK":"add","EDIT_TASK":"edit","DELETE_TASK":"delete","COMPLETE_TASK":"complete",
    "ASK_CLARIFICATION":"ask_clarification","ASK_CONFIRMATION":"confirm",
    "SHOW_TASKS":None,"DO_NOTHING":None,"UNDO":None,
}

def _parse_tools(raw: str) -> list[str | None]:
    normalized = raw.replace(" and "," or ")
    out = []
    for part in normalized.split(" or "):
        key = part.strip().split("(")[0].strip()
        if key in _A2T:
            out.append(_A2T[key])
    return out or [None]

def judge_turn(turn_result: dict, turn_spec: dict) -> dict:
    tool_names = [tc["name"] for tc in turn_result["tool_calls"]]
    must_not   = turn_spec.get("must_not", [])
    raw        = turn_spec["expected_action"]

    violations = [m for m in must_not if m in tool_names]
    if violations:
        return {"pass_fail":"FAIL","severe":True,
                "reasoning":f"Forbidden tool(s) called: {violations}. All: {tool_names or ['(none)']}"}

    if "SHOW_TASKS" in raw:
        return {"pass_fail":"PASS","severe":False,
                "reasoning":f"Read-only — no forbidden tools. Actual: {tool_names or ['(none)']}"}

    if "DO_NOTHING" in raw:
        if not tool_names:
            return {"pass_fail":"PASS","severe":False,"reasoning":"Correctly did nothing."}
        return {"pass_fail":"FAIL","severe":True,
                "reasoning":f"Expected DO_NOTHING but called: {tool_names}"}

    if "UNDO" in raw:
        if not tool_names or "edit" in tool_names:
            return {"pass_fail":"PASS","severe":False,
                    "reasoning":f"Acceptable undo response. Tools: {tool_names or ['(none)']}"}
        return {"pass_fail":"FAIL","severe":False,
                "reasoning":f"Expected UNDO behavior, got: {tool_names}"}

    expected = _parse_tools(raw)
    non_none  = [t for t in expected if t is not None]
    if any(t in tool_names for t in non_none):
        return {"pass_fail":"PASS","severe":False,
                "reasoning":f"Expected tool found. Actual: {tool_names}"}
    return {"pass_fail":"FAIL","severe":False,
            "reasoning":f"Expected one of {non_none} but got: {tool_names or ['(none)']}"}

# ---------------------------------------------------------------------------
# Per-test runner
# ---------------------------------------------------------------------------

def run_test(tc: dict) -> dict:
    runner = TodoAgentRunner(tc["initial_todos"])
    turn_results = []
    for turn_spec in tc["turns"]:
        result  = runner.run_turn(turn_spec["user"])
        verdict = judge_turn(result, turn_spec)
        turn_results.append({
            "user": turn_spec["user"],
            "expected_action": turn_spec["expected_action"],
            "tool_calls": result["tool_calls"],
            "response_snippet": (result["response"] or "")[:300],
            **verdict,
        })
    overall = "PASS" if all(t["pass_fail"] == "PASS" for t in turn_results) else "FAIL"
    return {"id":tc["id"],"category":tc["category"],"description":tc["description"],
            "turns":turn_results,"overall":overall,
            "severe":any(t.get("severe") for t in turn_results)}

# ---------------------------------------------------------------------------
# Single simulation runner
# ---------------------------------------------------------------------------

def run_simulation(sim_num: int) -> list[dict]:
    results: list[dict] = []
    bar = tqdm(
        total=len(TEST_CASES),
        desc=f"Sim {sim_num}/{SIMULATIONS}",
        unit="test",
        ncols=90,
        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
    )
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(run_test, tc): tc["id"] for tc in TEST_CASES}
        for future in concurrent.futures.as_completed(futures):
            tid = futures[future]
            try:
                r = future.result()
                results.append(r)
                badge = "✅" if r["overall"] == "PASS" else "❌"
                bar.set_postfix_str(f"{badge} {tid}", refresh=True)
            except Exception as exc:
                tqdm.write(f"  💥 [{sim_num}] {tid} — {exc}")
                results.append({"id":tid,"category":"?","description":"crashed",
                                 "turns":[],"overall":"ERROR","severe":False})
            finally:
                bar.update(1)
    bar.close()
    results.sort(key=lambda r: r["id"])
    return results

# ---------------------------------------------------------------------------
# Report printer
# ---------------------------------------------------------------------------
_W = 115

def print_sim_report(sim_num: int, results: list[dict], elapsed: float):
    passed = sum(1 for r in results if r["overall"] == "PASS")
    total  = len(results)
    acc    = 100 * passed / total if total else 0
    print(f"\n{'='*_W}")
    print(f"  SIMULATION {sim_num}  |  Score: {passed}/{total}  |  Accuracy: {acc:.1f}%  |  Time: {elapsed:.1f}s")
    print(f"{'='*_W}")
    print(f"{'ID':<6} {'CATEGORY':<28} {'RESULT':<10} {'SEV':<5} DESCRIPTION")
    print(f"{'-'*_W}")
    for r in results:
        badge = "✅ PASS" if r["overall"] == "PASS" else "❌ FAIL"
        sev   = "⚠" if r["severe"] else "—"
        print(f"{r['id']:<6} {r['category'][:27]:<28} {badge:<10} {sev:<5} {r['description']}")
        for i, t in enumerate(r["turns"], 1):
            calls = (", ".join(f"{c['name']}({list(c['args'].keys())})" for c in t["tool_calls"])
                     or "(none)")
            v = "✅" if t["pass_fail"] == "PASS" else "❌"
            print(f"         Turn {i} {v}  Expected: {t['expected_action']}")
            print(f"                Tools: {calls}")
            print(f"                Why:   {t['reasoning']}")
    print(f"{'='*_W}\n")

# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _cell_style(ws, row, col, value, fill=None, bold=False, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill = fill
    if bold:   cell.font = Font(bold=True)
    if wrap:   cell.alignment = Alignment(wrap_text=True, vertical="top")
    else:      cell.alignment = Alignment(vertical="top")
    return cell

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
HEADER = PatternFill("solid", fgColor="210235")

def export_excel(all_sim_results: list[dict]):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    headers = ["Simulation", "Total Tests", "Passed", "Failed", "Accuracy %", "Severe Errors", "Time (s)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, sim in enumerate(all_sim_results, 2):
        results = sim["results"]
        passed  = sum(1 for r in results if r["overall"] == "PASS")
        failed  = sum(1 for r in results if r["overall"] != "PASS")
        severe  = sum(1 for r in results if r["severe"])
        acc     = round(100 * passed / len(results), 1) if results else 0
        row_data = [sim["sim_num"], len(results), passed, failed, acc, severe, round(sim["elapsed"],1)]
        fill = GREEN if acc >= 80 else (YELLOW if acc >= 60 else RED)
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # Aggregate row
    agg_row = len(all_sim_results) + 2
    all_results = [r for sim in all_sim_results for r in sim["results"]]
    total_pass = sum(1 for r in all_results if r["overall"] == "PASS")
    agg_acc = round(100 * total_pass / len(all_results), 1) if all_results else 0
    ws.cell(row=agg_row, column=1, value="AGGREGATE").font = Font(bold=True)
    ws.cell(row=agg_row, column=2, value=len(all_results))
    ws.cell(row=agg_row, column=3, value=total_pass)
    ws.cell(row=agg_row, column=4, value=len(all_results) - total_pass)
    ws.cell(row=agg_row, column=5, value=agg_acc)

    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 2: All Results ──────────────────────────────────────────────
    ws2 = wb.create_sheet("All Results")
    h2 = ["Sim","Test ID","Category","Description","Turn","User Input",
          "Expected Action","Actual Tools Called","Agent Response","Pass/Fail","Severe"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HEADER; cell.font = Font(bold=True, color="FFFFFF")

    row2 = 2
    for sim in all_sim_results:
        for r in sim["results"]:
            for t_idx, t in enumerate(r["turns"], 1):
                calls = (", ".join(f"{c['name']}({json.dumps(c['args'],separators=(',',':'))})"
                                   for c in t["tool_calls"]) or "(none)")
                fill2 = GREEN if t["pass_fail"] == "PASS" else RED
                vals = [sim["sim_num"], r["id"], r["category"], r["description"], t_idx,
                        t["user"], t["expected_action"], calls,
                        t["response_snippet"], t["pass_fail"],
                        "YES" if t.get("severe") else "no"]
                for c, v in enumerate(vals, 1):
                    cell = ws2.cell(row=row2, column=c, value=v)
                    cell.fill = fill2
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                row2 += 1

    col_widths2 = [6, 8, 22, 35, 5, 45, 30, 50, 60, 8, 7]
    for c, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[1].height = 18

    # ── Sheet 3: Failures Only ────────────────────────────────────────────
    ws3 = wb.create_sheet("Failures")
    h3 = ["Sim","Test ID","Category","Turn","User Input (all turns)",
          "Expected Action","Actual Tools Called","Agent Response","Severe"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = HEADER; cell.font = Font(bold=True, color="FFFFFF")

    row3 = 2
    for sim in all_sim_results:
        for r in sim["results"]:
            failing_turns = [t for t in r["turns"] if t["pass_fail"] != "PASS"]
            if not failing_turns:
                continue
            all_user_inputs = " → ".join(t["user"] for t in r["turns"])
            for t in failing_turns:
                calls = (", ".join(f"{c['name']}({json.dumps(c['args'],separators=(',',':'))})"
                                   for c in t["tool_calls"]) or "(none)")
                vals = [sim["sim_num"], r["id"], r["category"],
                        r["turns"].index(t) + 1,
                        all_user_inputs, t["expected_action"], calls,
                        t["response_snippet"], "YES" if t.get("severe") else "no"]
                for c, v in enumerate(vals, 1):
                    cell = ws3.cell(row=row3, column=c, value=v)
                    cell.fill = RED if not t.get("severe") else YELLOW
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                row3 += 1

    col_widths3 = [6, 8, 22, 5, 55, 30, 55, 60, 7]
    for c, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    fname = f"eval_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    return fname

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------

def ensure_model():
    print(f"Verifying model '{MODEL}'...")
    try:
        ollama.pull(MODEL)
        print("  Model ready.\n")
    except Exception as exc:
        print(f"  Error: {exc}\n  Run: bash setup_local.sh")
        sys.exit(1)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 65)
    print("  Todo Agent — Evaluation Harness")
    print(f"  Model      : {MODEL}")
    print(f"  Test cases : {len(TEST_CASES)}")
    print(f"  Simulations: {SIMULATIONS}")
    print(f"  Workers    : {MAX_WORKERS} (parallel per simulation)")
    print("=" * 65 + "\n")

    ensure_model()

    # Rough time estimate
    est_per_sim = len(TEST_CASES) * 10 / MAX_WORKERS  # optimistic
    print(f"Estimated time: ~{SIMULATIONS * est_per_sim / 60:.0f}–{SIMULATIONS * len(TEST_CASES) * 10 / 60:.0f} min\n")

    all_sim_results = []
    for sim_num in range(1, SIMULATIONS + 1):
        print(f"\n{'─'*65}")
        print(f"  Starting Simulation {sim_num}/{SIMULATIONS} …")
        print(f"{'─'*65}")
        start = time.monotonic()
        results = run_simulation(sim_num)
        elapsed = time.monotonic() - start
        all_sim_results.append({"sim_num": sim_num, "results": results, "elapsed": elapsed})
        print_sim_report(sim_num, results, elapsed)

    # Aggregate summary
    print(f"\n{'='*65}")
    print("  AGGREGATE SUMMARY")
    print(f"{'='*65}")
    for sim in all_sim_results:
        passed = sum(1 for r in sim["results"] if r["overall"] == "PASS")
        total  = len(sim["results"])
        print(f"  Sim {sim['sim_num']}: {passed}/{total}  ({100*passed/total:.1f}%)  — {sim['elapsed']:.1f}s")

    all_r = [r for s in all_sim_results for r in s["results"]]
    total_pass = sum(1 for r in all_r if r["overall"] == "PASS")
    print(f"\n  Overall: {total_pass}/{len(all_r)}  ({100*total_pass/len(all_r):.1f}%)")

    severe_total = sum(1 for r in all_r if r["severe"])
    if severe_total:
        print(f"  Severe errors: {severe_total}")

    print(f"\nExporting results to Excel …")
    fname = export_excel(all_sim_results)
    print(f"  Saved: {fname}")
    print()

if __name__ == "__main__":
    main()
